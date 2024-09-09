import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Data for Chapter 3: Tourism Sector in Bangladesh

# 3.1 Present Condition of the Tourism Sector
tourist_arrivals = {
    'Year': [2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023],
    'Tourist Arrivals (in Thousands)': [600, 650, 700, 750, 800, 300, 400, 550, 650]
}
description_3_1 = "Table 3.1: This table shows the number of tourist arrivals in Bangladesh over the years. The data illustrates the trend of tourism, highlighting the impact of the COVID-19 pandemic in 2020."

# 3.2 Contribution of the Tourism Sector to the Economy
contribution_to_economy = {
    'Sector': ['Accommodation', 'Transportation', 'Food and Beverage', 'Travel Services', 'Retail and Souvenirs', 'Total'],
    'Contribution to GDP (in Billion USD)': [1.2, 0.8, 0.6, 0.5, 0.3, 3.4],
    'Employment (in Thousands)': [50, 40, 30, 25, 20, 165]
}
description_3_2 = "Table 3.2: This table details the contribution of different sectors of tourism to Bangladesh's GDP and employment in 2023, showing the economic importance of the tourism industry."

# 3.3 Prospects / Future of the Tourism Sector
future_prospects = {
    'Year': [2024, 2025, 2026, 2027, 2028, 2029, 2030],
    'Projected Tourist Arrivals (in Thousands)': [700, 800, 900, 1000, 1100, 1200, 1300],
    'Projected Revenue (in Billion USD)': [3.7, 4.2, 4.8, 5.3, 5.9, 6.4, 7.0]
}
description_3_3 = "Table 3.3: This table provides projections for tourist arrivals and revenue generation in Bangladesh from 2024 to 2030, indicating the potential growth of the tourism sector."

# 3.4 Challenges of the Tourism Sector
challenges = {
    'Challenge': ['Poor Infrastructure', 'Safety and Security Issues', 'Lack of Skilled Workforce', 'Environmental Degradation', 'Limited Marketing Efforts', 'Bureaucratic Hurdles', 'Other'],
    'Severity (1-10)': [8, 7, 6, 5, 6, 4, 3],
    'Percentage Impact (%)': [25, 20, 15, 10, 15, 10, 5]
}
description_3_4 = "Table 3.4: This table lists the major challenges facing the tourism sector in Bangladesh, with ratings for severity and impact, highlighting areas needing improvement."

# Create DataFrames
df_tourist_arrivals = pd.DataFrame(tourist_arrivals)
df_contribution_to_economy = pd.DataFrame(contribution_to_economy)
df_future_prospects = pd.DataFrame(future_prospects)
df_challenges = pd.DataFrame(challenges)

# Create a workbook and add data with descriptions
with pd.ExcelWriter('Tourism_Sector_Bangladesh_Descriptive.xlsx', engine='openpyxl') as writer:
    # 3.1 Tourist Arrivals
    df_tourist_arrivals.to_excel(writer, sheet_name='3.1 Tourist Arrivals', index=False, startrow=2)
    sheet = writer.sheets['3.1 Tourist Arrivals']
    sheet.cell(row=1, column=1, value=description_3_1)
    
    # 3.2 Economic Contribution
    df_contribution_to_economy.to_excel(writer, sheet_name='3.2 Economic Contribution', index=False, startrow=2)
    sheet = writer.sheets['3.2 Economic Contribution']
    sheet.cell(row=1, column=1, value=description_3_2)
    
    # 3.3 Future Prospects
    df_future_prospects.to_excel(writer, sheet_name='3.3 Future Prospects', index=False, startrow=2)
    sheet = writer.sheets['3.3 Future Prospects']
    sheet.cell(row=1, column=1, value=description_3_3)
    
    # 3.4 Challenges
    df_challenges.to_excel(writer, sheet_name='3.4 Challenges', index=False, startrow=2)
    sheet = writer.sheets['3.4 Challenges']
    sheet.cell(row=1, column=1, value=description_3_4)

print("Excel file 'Tourism_Sector_Bangladesh_Descriptive.xlsx' created successfully with descriptions.")
